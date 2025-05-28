import email
import imaplib
import os
import re
import smtplib
import time
from email.header import decode_header
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.utils import parseaddr
import openpyxl
from selenium import webdriver
from dotenv import load_dotenv
from mail.mail_config import MailConfig


class Mail:

    mailbox : MailConfig
    reply_subject = None

    def __init__(self, mailbox):
        self.mailbox = MailConfig(mailbox)

    def get_mail(self, filter_criteria):

        """функция получает почту по протоколу аймап,
        ищет в папке входящих, через filter_criteria - задаётся условие поиска, UNSEEN например - все непрочитанные
        далее берет это сообщение и поочерёдно декодирует имя отправителя, адрес отправителя, тему письма и тело письма
        если в теле письма обнаруживаются ссылки то автоматически происходит их вызов через селениум"""

        mail = imaplib.IMAP4_SSL("imap.mail.ru")  # подключаемся к серверу
        try:
            mail.login(self.mailbox.mail, self.mailbox.password)  # логинимся
            mail.select('inbox')  # выбираем папку, которую будем проверять

            status, messages = mail.search(None, filter_criteria)
            mail_ids = messages[0].split()

            # Получаем последнее письмо
            if mail_ids:
                last_email_id = mail_ids[-1]
                status, msg_data = mail.fetch(last_email_id, '(RFC822)')
                msg = email.message_from_bytes(msg_data[0][1])

                # Декодируем заголовок "От" и получаем имя отправителя
                sender_name = decode_header(msg['From'])[0][0]
                if isinstance(sender_name, bytes):
                    sender_name = sender_name.decode()
                sender_name = self.extract_sender_name(sender_name)

                # Получаем адрес отправителя
                sender_email = parseaddr(msg['From'])[1]

                # Декодируем заголовок "Тема"
                subject = decode_header(msg['Subject'])[0][0]
                if isinstance(subject, bytes):
                    subject = subject.decode()

                    # Декодируем тело сообщения
                    body = ""
                    if msg.is_multipart():
                        for part in msg.walk():
                            # Если часть - текстовая
                            if part.get_content_type() == "text/plain":
                                body = part.get_payload(decode=True).decode()
                                break  # Выходим из цикла после первого текстового содержимого
                    else:
                        # Если сообщение не многочастное
                        body = msg.get_payload(decode=True).decode()

                    # Находим ссылки в теле сообщения, очищаем их и делаем переход по ним
                    urls = re.findall(r'https://[^\s<>"]+', body)
                    cleaned_urls = [url.rstrip('"\'>)') for url in urls]
                    if cleaned_urls:
                        print("Найденные ссылки в письме:")
                        for url in cleaned_urls:
                            print(url)
                            options = webdriver.ChromeOptions()
                            options.page_load_strategy = 'eager'
                            driver = webdriver.Chrome(options = options)
                            driver.get(url)
                            time.sleep(2)
                            driver.quit()

                    print("Получено новое сообщение!")
                    print(f'От: {sender_email}, Тема: {subject}')
                    print(body)
                    print("\n")

                    return sender_email, subject, body

        except Exception as e:
            print("Ошибка:", e)

        finally:
            mail.logout()


    def send_mail(self, recipient, subject, message_body):

        """функция отправляет почту по протоколу SMTP"""

        message = MIMEMultipart()           # создали объект сообщения
        message['From'] = self.mailbox.mail
        message['To'] = recipient                # кому
        message['Subject'] = subject        # тема сообщения

        message.attach(MIMEText(message_body, 'plain'))     # в объект сообщения мы передаём сообщение,
        # обозначив что его формат будет plain, и это сообщение будет экземпляром класса MIMEText

        try:
            connect = smtplib.SMTP_SSL(self.mailbox.smtp_server, self.mailbox.smtp_port)
            # connect.starttls()        используется только с smtplib.SMTP, для незащищенного соединения, которое затем переводится в защищенное.
            connect.login(user = self.mailbox.mail, password = self.mailbox.password)
            connect.send_message(msg = message)
            connect.quit()
        except Exception as e:
            print("Ошибка отправки сообщения", e)



    def reply_mail(self, recipient, reply_subject, message):

        """функция создаёт ответное сообщение, добавляя Re в тему сообщения,
        вызывает функцию генерации тела ответа
        и вызывает функцию отправки сообщения получая отправителя в аргументах"""

        self.reply_subject = f"Re: {reply_subject}"
        self.message = self.generate_body_message(input_message=message)
        self.send_mail(recipient = recipient,
                       subject = self.reply_subject,
                       message_body = self.message)
        print(f"Ответное сообщение успешно отправлено - {self.message}")
        print("\n")


    def generate_body_message(self, input_message):

        """функция ходит в иммитатор БД - файл dialog_model ,берет каждое слово из столбца А и ищет его в теле сообщения
        если находит - то возвращает ответ из столбца Б, который соответствует записи поля А"""

        file_path = os.path.join(os.path.dirname(__file__), 'dialog_model.xlsx')
        workbook = openpyxl.load_workbook(file_path)

        sheet = workbook.active
        keywords = []

        for row in sheet.iter_rows(min_row=1, min_col=1, max_col=1, values_only=True):
            if row[0]:  # Проверяем, что значение не пустое
                keywords.append(str(row[0]).strip())
        for i in keywords:
            if i in input_message:
                index = keywords.index(i)
                answer = []
                for row in sheet.iter_rows(min_row=1, min_col=2, max_col=2, values_only=True):
                    answer.append(str(row[0]).strip())
                return answer[index]

        other_answer = 'Большое спасибо за информацию, в ближайшее время мы свяжемся с Вами'
        return other_answer


    def extract_sender_name(self, from_header: str) -> str:

        """вспомогательная функция "извлекатель" - получает имя отправителя очищаяя от лишнего"""

        match = re.search(r'<([^>]+)>', from_header)
        if match:
            return match.group(1)
        return from_header.strip()