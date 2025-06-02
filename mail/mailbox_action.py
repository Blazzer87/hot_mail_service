import email
import imaplib
import os
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import allure
import openpyxl
from mail.mail_message import MailMessage
from mail.mailbox_config import MailConfig


class Mailbox:

    mailbox : MailConfig


    def __init__(self, mailbox):
        self.mailbox = MailConfig(mailbox)
        self.last_email_id = None
        self.reply_subject = None
        self.message = None


    def get_mail(self, filter_criteria):

        """функция получает почту по протоколу аймап,
        ищет в папке входящих, через filter_criteria - задаётся условие поиска, UNSEEN например - все непрочитанные
        далее берет это сообщение и поочерёдно декодирует имя отправителя, адрес отправителя, тему письма и тело письма
        если в теле письма обнаруживаются ссылки то автоматически происходит их вызов через селениум"""

        with allure.step('Создаём объект IMAP4_SSL для чтения почты'):
            mail = imaplib.IMAP4_SSL("imap.mail.ru")  # подключаемся к серверу

        try:
            with allure.step('Передаём логин и пароль для чтения почты в объект mail'):
                mail.login(self.mailbox.mail, self.mailbox.password)  # логинимся

            with allure.step('Получаем папку inbox в объект mail'):
                mail.select('inbox')  # выбираем папку, которую будем проверять

            with allure.step('Ищем письма по заданным критериям'):
                status, messages = mail.search(None, filter_criteria)
                allure.attach(str(messages), 'Результаты поиска писем')  # Логируем результаты поиска
                mail_ids = messages[0].split()

            # Получаем последнее письмо
            if mail_ids:
                with allure.step('Получаем ID последнего письма'):
                    self.last_email_id = mail_ids[-1]
                    allure.attach(str(self.last_email_id), 'ID последнего письма')  # Логируем ID письма

                with allure.step('Получаем данные последнего письма'):
                    status, msg_data = mail.fetch(self.last_email_id, '(RFC822)')
                    msg = email.message_from_bytes(msg_data[0][1])
                    allure.attach(str(msg), 'Данные последнего письма')  # Логируем данные письма

                    self.message = MailMessage(msg)

                    with allure.step('Извлекаем информацию из письма'):
                        sender_mail = self.message.get_sender_mailbox()
                        allure.attach(sender_mail, 'Отправитель')  # Логируем отправителя

                        subject = self.message.get_mail_subject()
                        allure.attach(subject, 'Тема письма')  # Логируем тему письма

                        body = self.message.get_message_body()
                        allure.attach(body, 'Тело письма')  # Логируем тело письма

                        message_time, message_time_for_allure = self.message.get_time_message()
                        allure.attach(message_time_for_allure, 'Время сообщения')  # Логируем время сообщения

                        print("Получено новое сообщение!")
                        print(f'От: {sender_mail}, Тема: {subject}')
                        print(message_time)
                        print(body)

                        return sender_mail, subject, body, message_time

        except Exception as e:
            with allure.step('Обработка ошибки'):
                print("Ошибка:", e)
                allure.attach(str(e), 'Ошибка')  # Логируем текст ошибки

        finally:
            with allure.step('Выходим из почтового ящика'):
                mail.logout()


    def open_url_from_message(self):
        with allure.step('Проверяем наличие ссылок в теле письма'):
            self.message.find_url_in_message()


    def mark_mail_as_unread(self):

        """Функция устанавливает признак 'непрочитано' письму по его айдишнику."""

        with allure.step('Создаём объект IMAP4_SSL для работы с почтой'):
            mail = imaplib.IMAP4_SSL("imap.mail.ru")  # подключаемся к серверу

        try:
            with allure.step('Передаём логин и пароль для доступа к почте'):
                mail.login(self.mailbox.mail, self.mailbox.password)  # логинимся

            with allure.step('Выбираем папку inbox'):
                mail.select('inbox')  # выбираем папку, которую будем проверять

            with allure.step('Устанавливаем флаг "непрочитано" для письма'):
                mail.store(self.last_email_id, '-FLAGS', '\Seen')
                allure.attach(self.last_email_id.decode(), 'ID письма')  # Логируем ID письма
                print(f"Письмо с ID {self.last_email_id} помечено как непрочитанное.")

        except Exception as e:
            with allure.step('Обработка ошибки'):
                print("Ошибка:", e)
                allure.attach(str(e), 'Ошибка')  # Логируем текст ошибки

        finally:
            with allure.step('Выходим из почтового ящика'):
                mail.logout()



    def send_mail(self, recipient, subject, message_body):

        """Функция отправляет почту по протоколу SMTP"""

        message = MIMEMultipart()  # создаем объект сообщения
        message['From'] = self.mailbox.mail
        message['To'] = recipient  # кому
        message['Subject'] = subject  # тема сообщения

        message.attach(MIMEText(message_body, 'plain'))  # добавляем текст сообщения

        try:
            with allure.step('Создаём подключение к SMTP серверу'):
                connect = smtplib.SMTP_SSL(self.mailbox.smtp_server, self.mailbox.smtp_port)

            with allure.step('Выполняем логин на SMTP сервере'):
                connect.login(user=self.mailbox.mail, password=self.mailbox.password)

            with allure.step('Отправляем сообщение'):
                connect.send_message(msg=message)
                allure.attach(str(message), 'Отправляемое сообщение')  # Логируем отправляемое сообщение
                print(f"Ответное сообщение успешно отправлено - {message_body}")
                print("\n")

        except Exception as e:
            with allure.step('Обработка ошибки отправки сообщения'):
                print("Ошибка отправки сообщения:", e)
                allure.attach(str(e), 'Ошибка отправки')  # Логируем текст ошибки

        finally:
            with allure.step('Закрываем соединение с SMTP сервером'):
                connect.quit()



    def reply_mail(self, sender_mail, reply_subject, message):

        """функция создаёт ответное сообщение, добавляя Re в тему сообщения,
        вызывает функцию генерации тела ответа
        и вызывает функцию отправки сообщения получая отправителя в аргументах"""

        self.reply_subject = f"Re: {reply_subject}"
        self.message = self.generate_body_message(input_message=message)
        self.send_mail(recipient = sender_mail,
                       subject = self.reply_subject,
                       message_body = self.message)



    def generate_body_message(self, input_message: str):

        """Функция ходит в иммитатор БД - файл dialog_model, берет каждое слово из столбца А и ищет его в теле сообщения,
        если находит - то возвращает ответ из столбца Б, который соответствует записи поля А"""

        file_path = os.path.join(os.path.dirname(__file__), 'dialog_model.xlsx')
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        keywords = []

        with allure.step('Чтение ключевых слов из файла'):
            for row in sheet.iter_rows(min_row=1, min_col=1, max_col=1, values_only=True):
                if row[0]:  # Проверяем, что значение не пустое
                    keyword = str(row[0]).strip()
                    keywords.append(keyword)
                    allure.attach(keyword, 'Ключевое слово')  # Логируем каждое ключевое слово

        with allure.step('Поиск совпадений в сообщении'):
            for i in keywords:
                if i.lower() in input_message.lower():
                    index = keywords.index(i)
                    answer = []

                    with allure.step('Получение ответа из файла'):
                        for row in sheet.iter_rows(min_row=1, min_col=2, max_col=2, values_only=True):
                            answer.append(str(row[0]).strip())
                        response = answer[index]
                        allure.attach(response, 'Ответ')  # Логируем найденный ответ
                        return response

        other_answer = 'Большое спасибо за информацию, в ближайшее время мы свяжемся с Вами'
        with allure.step('Не найдено совпадений'):
            allure.attach(other_answer, 'Ответ по умолчанию')  # Логируем ответ по умолчанию
        return other_answer

